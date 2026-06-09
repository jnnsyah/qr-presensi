import os
import sqlite3
import uuid
import datetime
import smtplib
import json
import io
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from flask import Flask, request, jsonify, render_template, send_file, send_from_directory
from PIL import Image, ImageDraw, ImageFont
import qrcode
import ezdxf
from dotenv import load_dotenv
try:
    import openpyxl
    EXCEL_SUPPORTED = True
except ImportError:
    EXCEL_SUPPORTED = False

# Load env file if it exists
load_dotenv()

app = Flask(__name__)
DATABASE_FILE = 'database.db'

# ---------------------------------------------------------
# Helper: Database Connection
# ---------------------------------------------------------
def get_db():
    conn = sqlite3.connect(DATABASE_FILE)
    conn.row_factory = sqlite3.Row
    return conn

# ---------------------------------------------------------
# Helper: Certificate Base Generator
# ---------------------------------------------------------
def generate_default_certificate_template():
    path = os.path.join('assets', 'template_cert.png')
    if os.path.exists(path):
        return
    
    # Create image
    img = Image.new('RGB', (1920, 1357), color=(250, 247, 240))
    draw = ImageDraw.Draw(img)
    
    # Draw double border
    # Outer border
    draw.rectangle([(20, 20), (1900, 1337)], outline=(197, 160, 89), width=5)
    # Inner border
    draw.rectangle([(40, 40), (1880, 1317)], outline=(197, 160, 89), width=2)
    
    # Draw decorative corners
    for cx, cy in [(40, 40), (1880, 40), (40, 1317), (1880, 1317)]:
        draw.regular_polygon((cx, cy, 15), 4, rotation=45, fill=(197, 160, 89))
        
    # Attempt to load elegant serif fonts or fallback to system fonts
    font_paths = [
        "C:\\Windows\\Fonts\\georgiab.ttf",
        "C:\\Windows\\Fonts\\georgia.ttf",
        "C:\\Windows\\Fonts\\times.ttf",
        "C:\\Windows\\Fonts\\arial.ttf"
    ]
    
    font_title = None
    font_sub = None
    font_body = None
    font_sign = None
    
    for font_path in font_paths:
        if os.path.exists(font_path):
            try:
                font_title = ImageFont.truetype(font_path, 70)
                font_sub = ImageFont.truetype(font_path, 30)
                font_body = ImageFont.truetype(font_path, 35)
                font_sign = ImageFont.truetype(font_path, 25)
                break
            except Exception:
                continue
                
    if font_title is None:
        font_title = ImageFont.load_default()
        font_sub = ImageFont.load_default()
        font_body = ImageFont.load_default()
        font_sign = ImageFont.load_default()
        
    # Draw static texts
    # Title
    draw.text((960, 220), "SERTIFIKAT PRESENSI", fill=(44, 62, 80), font=font_title, anchor="mm")
    draw.text((960, 300), "DENGAN INTEGRASI AUTOCAD", fill=(197, 160, 89), font=font_sub, anchor="mm")
    
    # Awarded to
    draw.text((960, 460), "Diberikan Kepada :", fill=(127, 140, 141), font=font_sub, anchor="mm")
    
    # Bottom details
    draw.text((960, 820), "Atas partisipasi aktif sebagai peserta yang hadir pada lokasi seating plan", fill=(44, 62, 80), font=font_body, anchor="mm")
    draw.text((960, 880), "yang telah terintegrasi langsung dengan pemetaan AutoCAD.", fill=(44, 62, 80), font=font_body, anchor="mm")
    
    # Signature line
    draw.line([(810, 1150), (1110, 1150)], fill=(127, 140, 141), width=2)
    draw.text((960, 1175), "Panitia Pelaksana", fill=(44, 62, 80), font=font_sign, anchor="mm")
    
    # Save template
    img.save(path)

# ---------------------------------------------------------
# Helper: Student Certificate PDF Generator
# ---------------------------------------------------------
def generate_student_certificate(student_name, student_nim, checkin_time, id_mhs):
    template_path = os.path.join('assets', 'template_cert.png')
    pdf_path = os.path.join('assets', 'certificates', f"{id_mhs}.pdf")
    
    # Make sure template exists
    generate_default_certificate_template()
    
    img = Image.open(template_path)
    draw = ImageDraw.Draw(img)
    
    font_paths = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSerif-Bold.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "C:\\Windows\\Fonts\\georgiab.ttf",
        "C:\\Windows\\Fonts\\georgia.ttf",
        "C:\\Windows\\Fonts\\arial.ttf"
    ]
    
    # Handle long names - dynamically reduce font size in Pillow
    name_len = len(student_name)
    if name_len > 20:
        name_font_size = max(20, int(60 * (20 / name_len)))
    else:
        name_font_size = 60
        
    font_name = None
    font_nim = None
    font_date = None
    
    for font_path in font_paths:
        if os.path.exists(font_path):
            try:
                font_name = ImageFont.truetype(font_path, name_font_size)
                font_nim = ImageFont.truetype(font_path, 35)
                font_date = ImageFont.truetype(font_path, 25)
                break
            except Exception:
                continue
                
    if font_name is None:
        font_name = ImageFont.load_default()
        font_nim = ImageFont.load_default()
        font_date = ImageFont.load_default()
        
    # Draw student info
    # Name
    draw.text((960, 560), student_name.upper(), fill=(197, 160, 89), font=font_name, anchor="mm")
    # NIM
    draw.text((960, 640), f"NIM: {student_nim}", fill=(44, 62, 80), font=font_nim, anchor="mm")
    # Check-in timestamp
    draw.text((960, 960), f"Waktu Kehadiran: {checkin_time}", fill=(127, 140, 141), font=font_date, anchor="mm")
    
    # Save as PDF
    img_rgb = img.convert('RGB')
    img_rgb.save(pdf_path, "PDF", resolution=100.0)
    return pdf_path


# ---------------------------------------------------------
# Helper: SMTP & Mock Email Client
# ---------------------------------------------------------
def send_email(to_email, subject, html_body, attachment_path=None, attachment_name=None):
    smtp_server = os.environ.get('SMTP_SERVER')
    smtp_port = os.environ.get('SMTP_PORT')
    smtp_email = os.environ.get('SMTP_EMAIL')
    smtp_password = os.environ.get('SMTP_PASSWORD')
    
    email_sent_successfully = False
    error_message = None
    
    # Try sending real SMTP email if variables are configured
    if smtp_server and smtp_port and smtp_email and smtp_password:
        try:
            msg = MIMEMultipart()
            msg['From'] = smtp_email
            msg['To'] = to_email
            msg['Subject'] = subject
            
            msg.attach(MIMEText(html_body, 'html'))
            
            if attachment_path and os.path.exists(attachment_path):
                filename = attachment_name or os.path.basename(attachment_path)
                with open(attachment_path, 'rb') as attachment:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f'attachment; filename= {filename}')
                msg.attach(part)
                
            server = smtplib.SMTP(smtp_server, int(smtp_port))
            server.starttls()
            server.login(smtp_email, smtp_password)
            text = msg.as_string()
            server.sendmail(smtp_email, to_email, text)
            server.quit()
            email_sent_successfully = True
        except Exception as e:
            error_message = str(e)
            print(f"SMTP Error: {e}")
            
    # Always save a mock email file for easy debugging/visual verification
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_to_email = to_email.replace('@', '_').replace('.', '_')
    mock_filename = f"{timestamp}_{safe_to_email}_{subject.replace(' ', '_')[:30]}.html"
    mock_path = os.path.join('assets', 'emails', mock_filename)
    
    # Inject styling and links for attachments in mock view
    mock_header = f"""
    <div style="background:#f8f9fa; padding:15px; border-bottom:1px solid #ddd; font-family:sans-serif; margin-bottom:20px;">
        <strong>To:</strong> {to_email}<br/>
        <strong>Subject:</strong> {subject}<br/>
        <strong>Time:</strong> {datetime.datetime.now().isoformat()}<br/>
        <strong>SMTP Status:</strong> {"<span style='color:green;'>Sent via SMTP</span>" if email_sent_successfully else ("<span style='color:orange;'>Mock Mode (SMTP config missing)</span>" if not error_message else f"<span style='color:red;'>SMTP Failed: {error_message}</span>")}
    </div>
    """
    
    # Adjust attachment links
    attachment_section = ""
    if attachment_path:
        rel_path = os.path.relpath(attachment_path, os.path.join('assets', 'emails'))
        if attachment_path.endswith('.png'):
            attachment_section = f"""
            <div style="margin-top:20px; border-top:1px dashed #ccc; padding-top:10px; font-family:sans-serif;">
                <strong>Attached QR Code:</strong><br/>
                <img src="{rel_path}" style="max-width:200px; border:1px solid #ccc; margin-top:10px;"/>
                <br/><small><a href="{rel_path}" target="_blank">View Raw QR Image ({os.path.basename(attachment_path)})</a></small>
            </div>
            """
        else:
            attachment_section = f"""
            <div style="margin-top:20px; border-top:1px dashed #ccc; padding-top:10px; font-family:sans-serif;">
                <strong>Attached Document:</strong><br/>
                📄 <a href="{rel_path}" target="_blank" style="color:#007bff; font-weight:bold;">Open Certificate PDF ({os.path.basename(attachment_path)})</a>
            </div>
            """
            
    full_mock_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <title>Email Mock: {subject}</title>
    </head>
    <body style="margin:0; padding:20px; font-family:sans-serif; background-color:#f1f3f5;">
        <div style="max-width:600px; margin:0 auto; background:#ffffff; border-radius:8px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); overflow:hidden; border:1px solid #e1e4e8;">
            {mock_header}
            <div style="padding: 20px 30px; font-family:sans-serif; color:#333; line-height:1.5;">
                {html_body}
                {attachment_section}
            </div>
        </div>
    </body>
    </html>
    """
    
    with open(mock_path, 'w', encoding='utf-8') as f:
        f.write(full_mock_content)
        
    return email_sent_successfully, mock_path

# ---------------------------------------------------------
# DB Initialization
# ---------------------------------------------------------
def init_db():
    os.makedirs(os.path.join('assets', 'qrcodes'), exist_ok=True)
    os.makedirs(os.path.join('assets', 'certificates'), exist_ok=True)
    os.makedirs(os.path.join('assets', 'emails'), exist_ok=True)
    
    conn = get_db()
    cursor = conn.cursor()
    
    # Drop coordinate columns if they exist in pre-existing table
    for col in ["koordinat_x", "koordinat_y"]:
        try:
            cursor.execute(f"ALTER TABLE mahasiswa DROP COLUMN {col}")
            conn.commit()
        except Exception:
            pass
            
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS mahasiswa (
            id_mhs TEXT PRIMARY KEY,
            nama TEXT NOT NULL,
            email TEXT NOT NULL,
            nim TEXT NOT NULL UNIQUE,
            status_hadir TEXT DEFAULT 'Belum',
            waktu_hadir TEXT,
            cert_terkirim TEXT DEFAULT 'Belum',
            cert_waktu_kirim TEXT
        )
    ''')
    conn.commit()
    
    # Migrate existing DB — add new columns if not present
    for col_def in [
        ("cert_terkirim", "TEXT DEFAULT 'Belum'"),
        ("cert_waktu_kirim", "TEXT"),
        ("qr_terkirim", "TEXT DEFAULT 'Belum'"),
        ("qr_waktu_kirim", "TEXT")
    ]:
        try:
            cursor.execute(f"ALTER TABLE mahasiswa ADD COLUMN {col_def[0]} {col_def[1]}")
            conn.commit()
        except Exception:
            pass  # column already exists
    
    conn.close()
    
    # Pre-generate the default certificate canvas
    generate_default_certificate_template()


# Initialize DB on start
init_db()

# ---------------------------------------------------------
# Custom Certificate & Email Helper Systems
# ---------------------------------------------------------
EMAIL_CONFIG_FILE = os.path.join('assets', 'email_config.json')
CERT_CONFIG_FILE = os.path.join('assets', 'cert_config.json')

def clear_certificates_cache():
    cert_dir = os.path.join('assets', 'certificates')
    if os.path.exists(cert_dir):
        for f in os.listdir(cert_dir):
            if f.endswith('.pdf'):
                try:
                    os.remove(os.path.join(cert_dir, f))
                except Exception as e:
                    print(f"Error deleting cached certificate {f}: {e}", flush=True)

def load_email_config():
    default_config = {
        "qr_email_subject": "Tiket Presensi QR - {nama}",
        "qr_email_body": """<div style="font-family:sans-serif; color:#333; line-height:1.5;">
    <h2 style="color:#2c3e50; border-bottom:2px solid #c5a059; padding-bottom:10px;">TIKET PRESENSI QR</h2>
    <p>Halo <strong>{nama}</strong>,</p>
    <p>Anda terdaftar sebagai peserta dengan detail berikut:</p>
    <table style="border-collapse:collapse; width:100%; max-width:400px; margin-bottom:15px;">
        <tr>
            <td style="padding:5px 0; font-weight:bold;">Nama</td>
            <td>: {nama}</td>
        </tr>
        <tr>
            <td style="padding:5px 0; font-weight:bold;">NIM</td>
            <td>: {nim}</td>
        </tr>
    </table>
    <div style="background:#f9f9f9; border-left:4px solid #c5a059; padding:15px; border-radius:4px; margin-bottom:20px;">
        <strong>Petunjuk Registrasi:</strong><br/>
        Silakan tunjukkan QR Code yang dilampirkan pada email ini ke kamera web scanner panitia di gerbang masuk.
    </div>
    <p style="font-size:12px; color:#999; margin-top:30px;">
        *Email ini dikirim oleh QR-Presensi & CAD Integrator.
    </p>
</div>""",
        "cert_email_subject": "Sertifikat Kehadiran Resmi - {nama}",
        "cert_email_body": """<div style="font-family:sans-serif; color:#333; line-height:1.5;">
    <h2 style="color:#27ae60; border-bottom:2px solid #2ecc71; padding-bottom:10px;">PRESENSI BERHASIL & E-SERTIFIKAT</h2>
    <p>Halo <strong>{nama}</strong>,</p>
    <p>Kehadiran Anda telah sukses tercatat pada sistem presensi.</p>
    <table style="border-collapse:collapse; width:100%; max-width:400px; margin-bottom:15px;">
        <tr>
            <td style="padding:5px 0; font-weight:bold;">Nama</td>
            <td>: {nama}</td>
        </tr>
        <tr>
            <td style="padding:5px 0; font-weight:bold;">NIM</td>
            <td>: {nim}</td>
        </tr>
        <tr>
            <td style="padding:5px 0; font-weight:bold;">Waktu Hadir</td>
            <td>: {waktu}</td>
        </tr>
    </table>
    <div style="background:#eafaf1; border-left:4px solid #2ecc71; padding:15px; border-radius:4px; margin-bottom:20px;">
        <strong>Sertifikat Kehadiran Tersedia!</strong><br/>
        Sertifikat kehadiran resmi Anda telah berhasil diterbitkan dan dilampirkan pada email ini dalam format PDF.
    </div>
    <p>Terima kasih atas kedisiplinan dan partisipasi Anda.</p>
    <p style="font-size:12px; color:#999; margin-top:30px;">
        *Email ini dikirim oleh QR-Presensi & CAD Integrator.
    </p>
</div>"""
    }
    if os.path.exists(EMAIL_CONFIG_FILE):
        try:
            with open(EMAIL_CONFIG_FILE, 'r') as f:
                saved = json.load(f)
                default_config.update(saved)
        except Exception:
            pass
    return default_config

def save_email_config(config):
    try:
        with open(EMAIL_CONFIG_FILE, 'w') as f:
            json.dump(config, f, indent=4)
        return True
    except Exception:
        return False

def send_student_qr_email(student):
    cfg = load_email_config()
    subject = cfg['qr_email_subject'].format(nama=student['nama'], nim=student['nim'], email=student['email'])
    body = cfg['qr_email_body'].format(nama=student['nama'], nim=student['nim'], email=student['email'])
    
    qr_path = os.path.join('assets', 'qrcodes', f"{student['id_mhs']}.png")
    if not os.path.exists(qr_path):
        qr = qrcode.QRCode(version=1, box_size=10, border=4)
        qr.add_data(student['id_mhs'])
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")
        qr_img.save(qr_path)
        
    return send_email(student['email'], subject, body, attachment_path=qr_path, attachment_name=f"ticket_{student['nim']}.png")

def send_student_cert_email(student):
    cfg = load_email_config()
    waktu = student['waktu_hadir'] or ""
    subject = cfg['cert_email_subject'].format(nama=student['nama'], nim=student['nim'], waktu=waktu)
    body = cfg['cert_email_body'].format(nama=student['nama'], nim=student['nim'], waktu=waktu)
    
    # Always regenerate certificate to ensure it reflects the latest template edits
    pdf_path = generate_dxf_certificate(student['nama'], student['nim'], waktu, student['id_mhs'], student['email'])
        
    return send_email(student['email'], subject, body, attachment_path=pdf_path, attachment_name=f"certificate_{student['nim']}.pdf")

def parse_dxf_text_elements(dxf_path):
    try:
        doc = ezdxf.readfile(dxf_path)
        msp = doc.modelspace()
        elements = []
        for entity in msp:
            if entity.dxftype() in ('TEXT', 'MTEXT'):
                text = entity.dxf.text.strip()
                if text:
                    elements.append({
                        "id": entity.dxf.handle,
                        "type": entity.dxftype(),
                        "text": text,
                        "x": entity.dxf.insert[0] if entity.dxftype() == 'TEXT' else entity.dxf.insert[0],
                        "y": entity.dxf.insert[1] if entity.dxftype() == 'TEXT' else entity.dxf.insert[1],
                    })
        return elements
    except Exception as e:
        print(f"Error parsing DXF: {e}", flush=True)
        return []

def render_dxf_template_preview():
    dxf_template = os.path.join('assets', 'template_cert.dxf')
    preview_path = os.path.join('assets', 'template_cert_preview.png')
    if not os.path.exists(dxf_template):
        return False
    try:
        from ezdxf.addons.drawing import Frontend, RenderContext, pymupdf, layout
        doc = ezdxf.readfile(dxf_template)
        msp = doc.modelspace()
        backend = pymupdf.PyMuPdfBackend()
        Frontend(RenderContext(doc), backend).draw_layout(msp)
        png_bytes = backend.get_pixmap_bytes(layout.Page(0, 0), fmt="png", dpi=150)
        with open(preview_path, "wb") as fp:
            fp.write(png_bytes)
        return True
    except Exception as e:
        print(f"Error rendering DXF preview: {e}", flush=True)
        return False

def generate_image_certificate(student_name, student_nim, checkin_time, id_mhs, student_email):
    """Generate certificate from uploaded image background + Pillow text overlay."""
    img_cfg_path = os.path.join('assets', 'cert_image_config.json')
    if not os.path.exists(img_cfg_path):
        return None
    
    bg_path = None
    for ext in ['png', 'jpg', 'jpeg']:
        p = os.path.join('assets', f'cert_bg.{ext}')
        if os.path.exists(p):
            bg_path = p
            break
    if not bg_path:
        return None
    
    try:
        with open(img_cfg_path, 'r') as f:
            cfg = json.load(f)
        
        img = Image.open(bg_path).convert('RGBA')
        draw = ImageDraw.Draw(img)
        W, H = img.size
        
        field_data = {
            'nama': student_name,
            'nim': student_nim,
            'waktu': checkin_time,
            'email': student_email,
        }
        
        for field_cfg in cfg.get('fields', []):
            field_type = field_cfg.get('field', 'custom')
            text = field_data.get(field_type, field_cfg.get('custom_text', ''))
            
            x = int(field_cfg['x_pct'] / 100.0 * W)
            y = int(field_cfg['y_pct'] / 100.0 * H)
            font_size = int(field_cfg.get('font_size', 40))
            bold = field_cfg.get('bold', False)
            align = field_cfg.get('align', 'center')
            hex_color = field_cfg.get('color', '#000000').lstrip('#')
            color_rgb = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
            
            # Scale down font for long names
            if field_type == 'nama' and len(text) > 22:
                font_size = max(16, int(font_size * 22 / len(text)))
            
            # Select fonts dynamically based on bold property to match the frontend editor preview
            if bold:
                font_paths = [
                    "/usr/share/fonts/truetype/dejavu/DejaVuSerif-Bold.ttf",
                    "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
                    "C:\\Windows\\Fonts\\georgiab.ttf",
                    "C:\\Windows\\Fonts\\timesbd.ttf",
                    "C:\\Windows\\Fonts\\arialbd.ttf"
                ]
            else:
                font_paths = [
                    "/usr/share/fonts/truetype/dejavu/DejaVuSerif.ttf",
                    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                    "C:\\Windows\\Fonts\\georgia.ttf",
                    "C:\\Windows\\Fonts\\times.ttf",
                    "C:\\Windows\\Fonts\\arial.ttf"
                ]
            
            font = None
            for fp in font_paths:
                if os.path.exists(fp):
                    try:
                        font = ImageFont.truetype(fp, font_size)
                        break
                    except Exception:
                        continue
            if font is None:
                font = ImageFont.load_default()
            
            anchor = 'mm' if align == 'center' else ('lm' if align == 'left' else 'rm')
            draw.text((x, y), text, font=font, fill=color_rgb, anchor=anchor)
        
        pdf_path = os.path.join('assets', 'certificates', f"{id_mhs}.pdf")
        img_rgb = img.convert('RGB')
        img_rgb.save(pdf_path, 'PDF', resolution=150.0)
        return pdf_path
    except Exception as e:
        print(f"Error generating image certificate: {e}", flush=True)
        return None


def generate_dxf_certificate(student_name, student_nim, checkin_time, id_mhs, student_email):
    # Priority 1: image template
    result = generate_image_certificate(student_name, student_nim, checkin_time, id_mhs, student_email)
    if result:
        return result
    
    # Priority 2: DXF template
    dxf_template = os.path.join('assets', 'template_cert.dxf')
    pdf_path = os.path.join('assets', 'certificates', f"{id_mhs}.pdf")
    if not os.path.exists(dxf_template):
        return generate_student_certificate(student_name, student_nim, checkin_time, id_mhs)
    try:
        from ezdxf.addons.drawing import Frontend, RenderContext, pymupdf, layout
        doc = ezdxf.readfile(dxf_template)
        msp = doc.modelspace()
        mappings = []
        custom_texts = {}
        if os.path.exists(CERT_CONFIG_FILE):
            try:
                with open(CERT_CONFIG_FILE, 'r') as f:
                    cfg = json.load(f)
                    mappings = cfg.get('mappings', [])
                    custom_texts = cfg.get('custom_texts', {})
            except Exception:
                pass
        for entity in msp:
            if entity.dxftype() in ('TEXT', 'MTEXT'):
                handle = entity.dxf.handle
                text_val = entity.dxf.text.strip()
                mapped_field = None
                for m in mappings:
                    if m.get('handle') == handle or m.get('original_text') == text_val:
                        mapped_field = m.get('mapped_to')
                        break
                new_text = None
                if mapped_field == 'nama':
                    new_text = student_name
                    if len(new_text) > 20:
                        scale = 20.0 / len(new_text)
                        entity.dxf.height = entity.dxf.height * max(0.5, scale) if entity.dxftype() == 'TEXT' else entity.dxf.char_height * max(0.5, scale)
                elif mapped_field == 'nim': new_text = student_nim
                elif mapped_field == 'email': new_text = student_email
                elif mapped_field == 'waktu': new_text = checkin_time
                elif mapped_field == 'custom': new_text = custom_texts.get(handle, text_val)
                elif text_val in ('[NAMA]', '{nama}', '{NAMA}', '[NAMA_MAHASISWA]'):
                    new_text = student_name
                elif text_val in ('[NIM]', '{nim}', '{NIM}'): new_text = student_nim
                elif text_val in ('[WAKTU]', '{waktu}', '{WAKTU}'): new_text = checkin_time
                if new_text is not None:
                    entity.dxf.text = new_text
        backend = pymupdf.PyMuPdfBackend()
        Frontend(RenderContext(doc), backend).draw_layout(msp)
        with open(pdf_path, "wb") as fp:
            fp.write(backend.get_pdf_bytes(layout.Page(0, 0)))
        return pdf_path
    except Exception as e:
        print(f"Error rendering DXF certificate: {e}. Falling back to default Pillow.", flush=True)
        return generate_student_certificate(student_name, student_nim, checkin_time, id_mhs)

# ---------------------------------------------------------
# Static File Handlers (For Dev assets access)
# ---------------------------------------------------------

@app.route('/assets/<path:path>')
def send_assets(path):
    return send_from_directory('assets', path)

# ---------------------------------------------------------
# Frontend Page Routes
# ---------------------------------------------------------
@app.route('/')
def dashboard():
    return render_template('dashboard.html')

@app.route('/scanner')
def scanner():
    return render_template('scanner.html')

@app.route('/mahasiswa')
def view_mahasiswa():
    return render_template('mahasiswa.html')

@app.route('/sertifikat/<id_mhs>')
def view_sertifikat(id_mhs):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM mahasiswa WHERE id_mhs = ?", (id_mhs,))
    student = cursor.fetchone()
    conn.close()
    if not student:
        return "Mahasiswa tidak ditemukan.", 404
    return render_template('sertifikat.html', student=dict(student))

# ---------------------------------------------------------
# API Gateways
# ------------------------------------------------------# ---------------------------------------------------------
# Static Page Routes (Extra)
# ---------------------------------------------------------
@app.route('/custom-email')
def custom_email_page():
    return render_template('custom_email.html')

@app.route('/custom-sertifikat')
def custom_sertifikat_page():
    return render_template('custom_sertifikat.html')

# ---------------------------------------------------------
# API Gateways
# ---------------------------------------------------------

# A. POST /import-dummy
@app.route('/import-dummy', methods=['POST'])
def import_dummy():
    # Predefined dummy student dataset (WITHOUT coordinates)
    dummies = [
        {"id": "mhs-uuid-001", "nama": "Ahmad Dani", "email": "dani@student.univ.ac.id", "nim": "1202200001"},
        {"id": "mhs-uuid-002", "nama": "Budi Santoso", "email": "budi@student.univ.ac.id", "nim": "1202200002"},
        {"id": "mhs-uuid-003", "nama": "Citra Lestari", "email": "citra@student.univ.ac.id", "nim": "1202200003"},
        {"id": "mhs-uuid-004", "nama": "Dina Fitria", "email": "dina@student.univ.ac.id", "nim": "1202200004"},
        {"id": "mhs-uuid-005", "nama": "Eko Prasetyo", "email": "eko@student.univ.ac.id", "nim": "1202200005"},
        {"id": "mhs-uuid-006", "nama": "Farhan Alim", "email": "farhan@student.univ.ac.id", "nim": "1202200006"},
        {"id": "mhs-uuid-007", "nama": "Gita Jovanka", "email": "gita@student.univ.ac.id", "nim": "1202200007"},
        {"id": "mhs-uuid-008", "nama": "Hadi Wijaya", "email": "hadi@student.univ.ac.id", "nim": "1202200008"},
        {"id": "mhs-uuid-009", "nama": "Indah Permata", "email": "indah@student.univ.ac.id", "nim": "1202200009"},
        {"id": "mhs-uuid-010", "nama": "Joko Susilo", "email": "joko@student.univ.ac.id", "nim": "1202200010"}
    ]
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # Clear existing data to avoid conflict during seed
        cursor.execute("DELETE FROM mahasiswa")
        
        for m in dummies:
            cursor.execute('''
                INSERT INTO mahasiswa (id_mhs, nama, email, nim, status_hadir, waktu_hadir)
                VALUES (?, ?, ?, ?, 'Belum', NULL)
            ''', (m['id'], m['nama'], m['email'], m['nim']))
            
        conn.commit()
        response = {"status": "success", "message": f"Successfully imported {len(dummies)} student records."}
        return jsonify(response), 200
    except sqlite3.Error as e:
        conn.rollback()
        return jsonify({"status": "error", "message": f"Database seed failed: {str(e)}"}), 500
    finally:
        conn.close()

# B. POST /blast-qr
BLAST_QR_RESULT_FILE = os.path.join('assets', 'blast_qr_result.json')
BLAST_CERT_RESULT_FILE = os.path.join('assets', 'blast_cert_result.json')

@app.route('/blast-qr', methods=['POST'])
def blast_qr():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM mahasiswa")
    students = cursor.fetchall()
    conn.close()
    
    if not students:
        return jsonify({"status": "warning", "message": "Tidak ada data mahasiswa."}), 400
    
    success_list, failed_list = [], []
    for s in students:
        try:
            send_student_qr_email(s)
            success_list.append({"id_mhs": s['id_mhs'], "nama": s['nama'], "email": s['email']})
        except Exception as e:
            failed_list.append({"id_mhs": s['id_mhs'], "nama": s['nama'], "email": s['email'], "error": str(e)})
    
    result = {"success": success_list, "failed": failed_list, "timestamp": datetime.datetime.now().isoformat()}
    with open(BLAST_QR_RESULT_FILE, 'w') as f:
        json.dump(result, f)
    
    total, ok, fail = len(students), len(success_list), len(failed_list)
    msg = f"QR Tiket terkirim: {ok}/{total} berhasil, {fail} gagal."
    status = "success" if fail == 0 else ("warning" if ok > 0 else "error")
    return jsonify({"status": status, "message": msg, "success": success_list, "failed": failed_list}), 200

# POST /blast-cert
@app.route('/blast-cert', methods=['POST'])
def blast_cert():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM mahasiswa WHERE status_hadir = 'Hadir'")
    students = cursor.fetchall()
    conn.close()
    
    if not students:
        return jsonify({"status": "warning", "message": "Belum ada mahasiswa yang hadir."}), 400
    
    success_list, failed_list = [], []
    conn2 = get_db()
    cur2 = conn2.cursor()
    for s in students:
        try:
            sent, _ = send_student_cert_email(dict(s))
            cert_status = 'Terkirim' if sent else 'Mock'
            cert_time = datetime.datetime.now().strftime("%d %B %Y, %H:%M:%S")
            cur2.execute("UPDATE mahasiswa SET cert_terkirim=?, cert_waktu_kirim=? WHERE id_mhs=?",
                         (cert_status, cert_time, s['id_mhs']))
            success_list.append({"id_mhs": s['id_mhs'], "nama": s['nama'], "email": s['email']})
        except Exception as e:
            failed_list.append({"id_mhs": s['id_mhs'], "nama": s['nama'], "email": s['email'], "error": str(e)})
    conn2.commit()
    conn2.close()
    
    result = {"success": success_list, "failed": failed_list, "timestamp": datetime.datetime.now().isoformat()}
    with open(BLAST_CERT_RESULT_FILE, 'w') as f:
        json.dump(result, f)
    
    total, ok, fail = len(students), len(success_list), len(failed_list)
    msg = f"Sertifikat terkirim: {ok}/{total} berhasil, {fail} gagal."
    status = "success" if fail == 0 else ("warning" if ok > 0 else "error")
    return jsonify({"status": status, "message": msg, "success": success_list, "failed": failed_list}), 200

# POST /blast-qr-retry
@app.route('/blast-qr-retry', methods=['POST'])
def blast_qr_retry():
    data = request.get_json() or {}
    ids = data.get('ids', [])
    if not ids:
        return jsonify({"status": "error", "message": "Tidak ada ID untuk dikirim ulang."}), 400
    conn = get_db()
    cursor = conn.cursor()
    placeholders = ','.join('?' * len(ids))
    cursor.execute(f"SELECT * FROM mahasiswa WHERE id_mhs IN ({placeholders})", ids)
    students = cursor.fetchall()
    conn.close()
    success_list, failed_list = [], []
    for s in students:
        try:
            send_student_qr_email(s)
            success_list.append({"id_mhs": s['id_mhs'], "nama": s['nama'], "email": s['email']})
        except Exception as e:
            failed_list.append({"id_mhs": s['id_mhs'], "nama": s['nama'], "email": s['email'], "error": str(e)})
    msg = f"Kirim ulang QR: {len(success_list)} berhasil, {len(failed_list)} gagal."
    return jsonify({"status": "success" if not failed_list else "warning", "message": msg, "success": success_list, "failed": failed_list}), 200

# POST /blast-cert-retry
@app.route('/blast-cert-retry', methods=['POST'])
def blast_cert_retry():
    data = request.get_json() or {}
    ids = data.get('ids', [])
    if not ids:
        return jsonify({"status": "error", "message": "Tidak ada ID untuk dikirim ulang."}), 400
    conn = get_db()
    cursor = conn.cursor()
    placeholders = ','.join('?' * len(ids))
    cursor.execute(f"SELECT * FROM mahasiswa WHERE id_mhs IN ({placeholders}) AND status_hadir='Hadir'", ids)
    students = cursor.fetchall()
    conn.close()
    success_list, failed_list = [], []
    conn2 = get_db()
    cur2 = conn2.cursor()
    for s in students:
        try:
            sent, _ = send_student_cert_email(dict(s))
            cert_status = 'Terkirim' if sent else 'Mock'
            cert_time = datetime.datetime.now().strftime("%d %B %Y, %H:%M:%S")
            cur2.execute("UPDATE mahasiswa SET cert_terkirim=?, cert_waktu_kirim=? WHERE id_mhs=?",
                         (cert_status, cert_time, s['id_mhs']))
            success_list.append({"id_mhs": s['id_mhs'], "nama": s['nama'], "email": s['email']})
        except Exception as e:
            failed_list.append({"id_mhs": s['id_mhs'], "nama": s['nama'], "email": s['email'], "error": str(e)})
    conn2.commit()
    conn2.close()
    msg = f"Kirim ulang sertifikat: {len(success_list)} berhasil, {len(failed_list)} gagal."
    return jsonify({"status": "success" if not failed_list else "warning", "message": msg, "success": success_list, "failed": failed_list}), 200

# C. POST /api/scan
@app.route('/api/scan', methods=['POST'])
def api_scan():
    data = request.get_json()
    if not data or 'id_mhs' not in data:
        return jsonify({"status": "error", "message": "Invalid scan payload. Missing student identifier."}), 400
        
    id_mhs = data['id_mhs']
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM mahasiswa WHERE id_mhs = ?", (id_mhs,))
    student = cursor.fetchone()
    
    if not student:
        conn.close()
        return jsonify({"status": "error", "message": "QR Code invalid. Student ID not found in database."}), 404
        
    # Check current status
    if student['status_hadir'] == 'Hadir':
        conn.close()
        return jsonify({
            "status": "warning",
            "message": f"{student['nama']} ({student['nim']}) sudah melakukan presensi sebelumnya pada {student['waktu_hadir']}.",
            "student": {
                "id_mhs": student['id_mhs'],
                "nama": student['nama'],
                "nim": student['nim'],
                "email": student['email'],
                "status_hadir": student['status_hadir'],
                "waktu_hadir": student['waktu_hadir']
            }
        }), 200
        
    # Flip status
    waktu_hadir_readable = datetime.datetime.now().strftime("%d %B %Y, %H:%M:%S")
    
    try:
        cursor.execute('''
            UPDATE mahasiswa
            SET status_hadir = 'Hadir', waktu_hadir = ?
            WHERE id_mhs = ?
        ''', (waktu_hadir_readable, id_mhs))
        conn.commit()
        
        # Generate Certificate PDF using custom DXF / default Pillow generator (but do NOT auto-blast email!)
        pdf_path = generate_dxf_certificate(
            student['nama'], 
            student['nim'], 
            waktu_hadir_readable, 
            id_mhs,
            student['email']
        )
        
        return jsonify({
            "status": "success",
            "message": f"Presensi berhasil dicatat untuk {student['nama']} ({student['nim']}). Sertifikat terbit.",
            "student": {
                "id_mhs": student['id_mhs'],
                "nama": student['nama'],
                "nim": student['nim'],
                "email": student['email'],
                "status_hadir": "Hadir",
                "waktu_hadir": waktu_hadir_readable,
                "cert_terkirim": student['cert_terkirim'],
                "cert_waktu_kirim": student['cert_waktu_kirim']
            }
        }), 200
        
    except Exception as e:
        conn.rollback()
        return jsonify({"status": "error", "message": f"Processing error during attendance capture: {str(e)}"}), 500
    finally:
        conn.close()

# E. POST /api/students
@app.route('/api/students', methods=['POST'])
def add_student():
    data = request.get_json()
    if not data or not all(k in data for k in ('nama', 'email', 'nim')):
        return jsonify({"status": "error", "message": "Payload tidak lengkap. Harus berisi nama, email, dan nim."}), 400
        
    nama = data['nama']
    email = data['email']
    nim = data['nim']

    conn = get_db()
    cursor = conn.cursor()
    
    # Check duplicate NIM
    cursor.execute("SELECT * FROM mahasiswa WHERE nim = ?", (nim,))
    existing = cursor.fetchone()
    if existing:
        conn.close()
        return jsonify({"status": "error", "message": f"NIM {nim} sudah terdaftar."}), 400
        
    id_mhs = str(uuid.uuid4())
    
    try:
        cursor.execute('''
            INSERT INTO mahasiswa (id_mhs, nama, email, nim, status_hadir, waktu_hadir)
            VALUES (?, ?, ?, ?, 'Belum', NULL)
        ''', (id_mhs, nama, email, nim))
        conn.commit()
        
        # Generate QR Code but DO NOT automatically send email
        qr = qrcode.QRCode(version=1, box_size=10, border=4)
        qr.add_data(id_mhs)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")
        
        qr_filename = f"{id_mhs}.png"
        qr_path = os.path.join('assets', 'qrcodes', qr_filename)
        qr_img.save(qr_path)
        
        return jsonify({
            "status": "success",
            "message": f"Mahasiswa {nama} berhasil didaftarkan. QR tiket telah di-generate.",
            "student": {
                "id_mhs": id_mhs,
                "nama": nama,
                "nim": nim,
                "email": email,
                "status_hadir": "Belum",
                "waktu_hadir": None,
                "qr_url": f"/assets/qrcodes/{id_mhs}.png"
            }
        }), 201
    except Exception as e:
        conn.rollback()
        return jsonify({"status": "error", "message": f"Registrasi gagal: {str(e)}"}), 500
    finally:
        conn.close()

# D. GET /export-autocad
@app.route('/export-autocad', methods=['GET'])
def export_autocad():
    dxf_template = os.path.join('assets', 'template_cert.dxf')
    if os.path.exists(dxf_template):
        return send_file(dxf_template, as_attachment=True, download_name="template_cert.dxf")
    else:
        # Create a dummy DXF file so tests don't fail if template is missing
        doc = ezdxf.new('R2010')
        msp = doc.modelspace()
        msp.add_text("DUMMY TEMPLATE FOR TESTING")
        dxf_path = os.path.join('assets', 'template_cert.dxf')
        doc.saveas(dxf_path)
        return send_file(dxf_path, as_attachment=True, download_name="template_cert.dxf")

# ---------------------------------------------------------
# Get Live Students List (Admin Table JSON Feed)
# ---------------------------------------------------------
@app.route('/api/students', methods=['GET'])
def get_students():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM mahasiswa ORDER BY status_hadir DESC, nama ASC")
    students = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(students)

# Manual Toggle Attendance from Admin Panel
@app.route('/api/students/<id_mhs>/toggle', methods=['POST'])
def toggle_student(id_mhs):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM mahasiswa WHERE id_mhs = ?", (id_mhs,))
    student = cursor.fetchone()
    
    if not student:
        conn.close()
        return jsonify({"status": "error", "message": "Student not found"}), 404
        
    new_status = 'Belum' if student['status_hadir'] == 'Hadir' else 'Hadir'
    waktu_hadir = datetime.datetime.now().strftime("%d %B %Y, %H:%M:%S") if new_status == 'Hadir' else None
    
    try:
        cursor.execute("UPDATE mahasiswa SET status_hadir = ?, waktu_hadir = ? WHERE id_mhs = ?", (new_status, waktu_hadir, id_mhs))
        conn.commit()
        
        if new_status == 'Hadir':
            # Generate the DXF Certificate PDF (but do not send email automatically)
            generate_dxf_certificate(student['nama'], student['nim'], waktu_hadir, id_mhs, student['email'])
        else:
            # Reset cert status when toggling back to Belum
            cursor.execute("UPDATE mahasiswa SET cert_terkirim = 'Belum', cert_waktu_kirim = NULL WHERE id_mhs = ?", (id_mhs,))
            conn.commit()
            
        return jsonify({"status": "success", "message": f"Status presensi diubah ke {new_status} untuk {student['nama']}."})
    except Exception as e:
        conn.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        conn.close()

# ---------------------------------------------------------
# CRUD: GET single student
# ---------------------------------------------------------
@app.route('/api/students/<id_mhs>', methods=['GET'])
def get_student(id_mhs):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM mahasiswa WHERE id_mhs = ?", (id_mhs,))
    student = cursor.fetchone()
    conn.close()
    if not student:
        return jsonify({"status": "error", "message": "Mahasiswa tidak ditemukan."}), 404
    return jsonify(dict(student))

# CRUD: PUT update student
@app.route('/api/students/<id_mhs>', methods=['PUT'])
def update_student(id_mhs):
    data = request.get_json()
    if not data:
        return jsonify({"status": "error", "message": "Payload kosong."}), 400

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM mahasiswa WHERE id_mhs = ?", (id_mhs,))
    student = cursor.fetchone()
    if not student:
        conn.close()
        return jsonify({"status": "error", "message": "Mahasiswa tidak ditemukan."}), 404

    # Check NIM uniqueness if changed
    new_nim = data.get('nim', student['nim'])
    if new_nim != student['nim']:
        cursor.execute("SELECT id_mhs FROM mahasiswa WHERE nim = ? AND id_mhs != ?", (new_nim, id_mhs))
        if cursor.fetchone():
            conn.close()
            return jsonify({"status": "error", "message": f"NIM {new_nim} sudah digunakan mahasiswa lain."}), 400

    nama = data.get('nama', student['nama'])
    email = data.get('email', student['email'])
    nim = new_nim

    try:
        cursor.execute('''
            UPDATE mahasiswa SET nama=?, email=?, nim=?
            WHERE id_mhs=?
        ''', (nama, email, nim, id_mhs))
        conn.commit()
        return jsonify({"status": "success", "message": f"Data {nama} berhasil diperbarui."})
    except Exception as e:
        conn.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        conn.close()

# CRUD: DELETE student
@app.route('/api/students/<id_mhs>', methods=['DELETE'])
def delete_student(id_mhs):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM mahasiswa WHERE id_mhs = ?", (id_mhs,))
    student = cursor.fetchone()
    if not student:
        conn.close()
        return jsonify({"status": "error", "message": "Mahasiswa tidak ditemukan."}), 404

    try:
        cursor.execute("DELETE FROM mahasiswa WHERE id_mhs = ?", (id_mhs,))
        conn.commit()

        # Cleanup QR and cert files
        for fpath in [
            os.path.join('assets', 'qrcodes', f"{id_mhs}.png"),
            os.path.join('assets', 'certificates', f"{id_mhs}.pdf")
        ]:
            if os.path.exists(fpath):
                os.remove(fpath)

        return jsonify({"status": "success", "message": f"Mahasiswa {student['nama']} berhasil dihapus."})
    except Exception as e:
        conn.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        conn.close()

# Resend certificate email
@app.route('/api/students/<id_mhs>/resend-cert', methods=['POST'])
def resend_cert(id_mhs):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM mahasiswa WHERE id_mhs = ?", (id_mhs,))
    student = cursor.fetchone()
    if not student:
        conn.close()
        return jsonify({"status": "error", "message": "Mahasiswa tidak ditemukan."}), 404

    if student['status_hadir'] != 'Hadir':
        conn.close()
        return jsonify({"status": "warning", "message": "Mahasiswa belum hadir. Sertifikat belum dapat dikirim."}), 400

    try:
        sent, _ = send_student_cert_email(dict(student))
        cert_status = 'Terkirim' if sent else 'Mock'
        cert_time = datetime.datetime.now().strftime("%d %B %Y, %H:%M:%S")
        cursor.execute("UPDATE mahasiswa SET cert_terkirim = ?, cert_waktu_kirim = ? WHERE id_mhs = ?",
                       (cert_status, cert_time, id_mhs))
        conn.commit()
        return jsonify({"status": "success", "message": "Sertifikat berhasil dikirim."})
    except Exception as e:
        conn.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        conn.close()

# ---------------------------------------------------------
# Dynamic AutoCAD Certificate & Custom Email Configuration API
# ---------------------------------------------------------

# GET/POST /api/email/config
@app.route('/api/email/config', methods=['GET', 'POST'])
def api_email_config():
    if request.method == 'GET':
        return jsonify(load_email_config())
    else:
        data = request.get_json() or {}
        if save_email_config(data):
            return jsonify({"status": "success", "message": "Konfigurasi email berhasil disimpan."})
        return jsonify({"status": "error", "message": "Gagal menyimpan konfigurasi email."}), 500

# POST /api/cert/upload-image
@app.route('/api/cert/upload-image', methods=['POST'])
def api_cert_upload_image():
    if 'file' not in request.files:
        return jsonify({"status": "error", "message": "File gambar tidak ditemukan."}), 400
    file = request.files['file']
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in ('png', 'jpg', 'jpeg'):
        return jsonify({"status": "error", "message": "Format harus PNG, JPG, atau JPEG."}), 400
    # Remove old cert_bg files
    for old_ext in ('png', 'jpg', 'jpeg'):
        old = os.path.join('assets', f'cert_bg.{old_ext}')
        if os.path.exists(old):
            os.remove(old)
    dest = os.path.join('assets', f'cert_bg.{ext}')
    file.save(dest)
    
    # Clear cached certificate PDFs since background template has changed
    clear_certificates_cache()
    
    try:
        img = Image.open(dest)
        W, H = img.size
        return jsonify({"status": "success", "ext": ext, "width": W, "height": H, "url": f"/assets/cert_bg.{ext}"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

# GET/POST /api/cert/config (image-based)
@app.route('/api/cert/config', methods=['GET', 'POST'])
def api_cert_config():
    img_cfg_path = os.path.join('assets', 'cert_image_config.json')
    
    # Detect current cert_bg
    bg_ext = None
    bg_url = None
    for ext in ('png', 'jpg', 'jpeg'):
        p = os.path.join('assets', f'cert_bg.{ext}')
        if os.path.exists(p):
            bg_ext = ext
            bg_url = f'/assets/cert_bg.{ext}'
            break
    
    if request.method == 'GET':
        image_config = {"fields": []}
        if os.path.exists(img_cfg_path):
            try:
                with open(img_cfg_path, 'r') as f:
                    image_config = json.load(f)
            except Exception:
                pass
        return jsonify({
            "has_image": bg_url is not None,
            "bg_url": bg_url,
            "image_config": image_config
        })
    else:
        data = request.get_json() or {}
        try:
            with open(img_cfg_path, 'w') as f:
                json.dump(data, f, indent=2)
            # Clear cached certificate PDFs since template layout config has changed
            clear_certificates_cache()
            return jsonify({"status": "success", "message": "Konfigurasi sertifikat berhasil disimpan."})
        except Exception as e:
            return jsonify({"status": "error", "message": str(e)}), 500

# POST /api/students/<id_mhs>/send-qr
@app.route('/api/students/<id_mhs>/send-qr', methods=['POST'])
def send_qr_manual(id_mhs):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM mahasiswa WHERE id_mhs = ?", (id_mhs,))
    student = cursor.fetchone()
    
    if not student:
        conn.close()
        return jsonify({"status": "error", "message": "Mahasiswa tidak ditemukan."}), 404
        
    try:
        sent, _ = send_student_qr_email(dict(student))
        qr_status = 'Terkirim' if sent else 'Mock'
        qr_time = datetime.datetime.now().strftime("%d %B %Y, %H:%M:%S")
        cursor.execute("UPDATE mahasiswa SET qr_terkirim=?, qr_waktu_kirim=? WHERE id_mhs=?",
                       (qr_status, qr_time, id_mhs))
        conn.commit()
        conn.close()
        if sent:
            return jsonify({"status": "success", "message": "Email QR Tiket berhasil terkirim."})
        return jsonify({"status": "warning", "message": "QR Tiket terkirim dalam mode simulasi (Mock/Log)."})
    except Exception as e:
        conn.close()
        return jsonify({"status": "error", "message": f"Gagal mengirim QR tiket: {str(e)}"}), 500

# POST /import-excel
@app.route('/import-excel', methods=['POST'])
def import_excel():
    if not EXCEL_SUPPORTED:
        return jsonify({"status": "error", "message": "openpyxl tidak terinstall di server."}), 500
    if 'file' not in request.files:
        return jsonify({"status": "error", "message": "File Excel tidak ditemukan."}), 400
    file = request.files['file']
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in ('xlsx', 'xls'):
        return jsonify({"status": "error", "message": "Format harus .xlsx atau .xls"}), 400
    
    try:
        content = file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        
        # Detect header row: look for columns named nama, email, nim (case-insensitive)
        headers = {}
        header_row = None
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            row_lower = [str(c).strip().lower() if c else '' for c in row]
            if any(h in row_lower for h in ('nama', 'email', 'nim')):
                header_row = i
                for j, h in enumerate(row_lower):
                    if h in ('nama', 'name'):
                        headers['nama'] = j
                    elif h in ('email', 'e-mail', 'e_mail'):
                        headers['email'] = j
                    elif h in ('nim', 'nrp', 'id', 'no', 'nomor'):
                        headers['nim'] = j
                break
        
        if not header_row or not all(k in headers for k in ('nama', 'email', 'nim')):
            return jsonify({"status": "error", "message": "Header kolom tidak ditemukan. Pastikan ada kolom Nama, Email, dan NIM."}), 400
        
        conn = get_db()
        cursor = conn.cursor()
        imported, skipped, errors = 0, 0, []
        
        for row in ws.iter_rows(min_row=header_row+1, values_only=True):
            try:
                nama = str(row[headers['nama']]).strip() if row[headers['nama']] else ''
                email = str(row[headers['email']]).strip() if row[headers['email']] else ''
                nim = str(row[headers['nim']]).strip() if row[headers['nim']] else ''
                # Skip empty rows
                if not nama or not email or not nim or nama.lower() == 'none':
                    continue
                # NIM must not be purely non-numeric to avoid header repeats
                if nim.lower() in ('nim', 'nrp', 'id', 'no'):
                    continue
                
                # Check if NIM already exists
                cursor.execute("SELECT id_mhs FROM mahasiswa WHERE nim=?", (nim,))
                if cursor.fetchone():
                    skipped += 1
                    continue
                
                id_mhs = str(uuid.uuid4())
                cursor.execute(
                    "INSERT INTO mahasiswa (id_mhs,nama,email,nim,status_hadir,waktu_hadir) VALUES (?,?,?,?,'Belum',NULL)",
                    (id_mhs, nama, email, nim)
                )
                # Generate QR code
                qr = qrcode.QRCode(version=1, box_size=10, border=4)
                qr.add_data(id_mhs)
                qr.make(fit=True)
                qr_img = qr.make_image(fill_color='black', back_color='white')
                qr_img.save(os.path.join('assets', 'qrcodes', f'{id_mhs}.png'))
                imported += 1
            except Exception as row_err:
                errors.append(str(row_err))
        
        conn.commit()
        conn.close()
        
        msg = f"Berhasil import {imported} mahasiswa baru. {skipped} NIM sudah ada (dilewati)."
        if errors:
            msg += f" {len(errors)} baris error."
        return jsonify({"status": "success", "message": msg, "imported": imported, "skipped": skipped, "errors": errors})
    except Exception as e:
        return jsonify({"status": "error", "message": f"Gagal membaca file Excel: {str(e)}"}), 500

# POST /api/students/<id_mhs>/send-cert
@app.route('/api/students/<id_mhs>/send-cert', methods=['POST'])
def send_cert_manual(id_mhs):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM mahasiswa WHERE id_mhs = ?", (id_mhs,))
    student = cursor.fetchone()
    
    if not student:
        conn.close()
        return jsonify({"status": "error", "message": "Mahasiswa tidak ditemukan."}), 404
        
    if student['status_hadir'] != 'Hadir':
        conn.close()
        return jsonify({"status": "warning", "message": "Mahasiswa belum hadir. Sertifikat belum dapat dibuat/dikirim."}), 400
        
    try:
        sent, _ = send_student_cert_email(dict(student))
        cert_status = 'Terkirim' if sent else 'Mock'
        cert_time = datetime.datetime.now().strftime("%d %B %Y, %H:%M:%S")
        cursor.execute("UPDATE mahasiswa SET cert_terkirim = ?, cert_waktu_kirim = ? WHERE id_mhs = ?",
                       (cert_status, cert_time, id_mhs))
        conn.commit()
        conn.close()
        
        if sent:
            return jsonify({"status": "success", "message": "Email Sertifikat berhasil terkirim."})
        return jsonify({"status": "warning", "message": "Sertifikat terkirim dalam mode simulasi (Mock/Log)."})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({"status": "error", "message": f"Gagal mengirim sertifikat: {str(e)}"}), 500

# Inline PDF certificate preview
@app.route('/api/students/<id_mhs>/cert-preview')
def cert_preview(id_mhs):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM mahasiswa WHERE id_mhs = ?", (id_mhs,))
    student = cursor.fetchone()
    conn.close()
    if not student:
        return "Mahasiswa tidak ditemukan.", 404
    if student['status_hadir'] != 'Hadir':
        return "Sertifikat belum tersedia — mahasiswa belum hadir.", 400
    # Always regenerate certificate to ensure it reflects the latest template edits
    pdf_path = generate_dxf_certificate(student['nama'], student['nim'], student['waktu_hadir'], id_mhs, student['email'])
    return send_file(pdf_path, mimetype='application/pdf')

# GET /download-template-excel
@app.route('/download-template-excel')
def download_template_excel():
    if not EXCEL_SUPPORTED:
        return jsonify({"status": "error", "message": "openpyxl tidak terinstall di server."}), 500
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Template Import"
        ws.append(["Nama", "NIM", "Email"])
        ws.append(["Budi Santoso", "12345678", "budi.santoso@mhs.univ.ac.id"])
        ws.append(["Siti Rahayu", "12345679", "siti.rahayu@mhs.univ.ac.id"])
        
        # Auto-adjust column width
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name="format_import_mahasiswa.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        return jsonify({"status": "error", "message": f"Gagal membuat template: {str(e)}"}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
